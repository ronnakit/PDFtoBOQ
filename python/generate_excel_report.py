# -*- coding: utf-8 -*-
import os
import sys
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import argparse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_boq(project_name='116-69 - แบบบ้านชั้นเดียว'):
    wb = openpyxl.Workbook()

    font_title = Font(name='Tahoma', size=14, bold=True, color='0F172A')
    font_header = Font(name='Tahoma', size=10, bold=True, color='FFFFFF')
    font_sub = Font(name='Tahoma', size=9, bold=True, color='334155')
    font_body = Font(name='Tahoma', size=9, color='1E293B')
    font_bold = Font(name='Tahoma', size=9, bold=True, color='0F172A')

    fill_header_orange = PatternFill(start_color='EA580C', end_color='EA580C', fill_type='solid')
    fill_header_dark = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    fill_total = PatternFill(start_color='FED7AA', end_color='FED7AA', fill_type='solid')
    fill_alt = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

    thin = Side(border_style='thin', color='CBD5E1')
    border_cell = Border(left=thin, right=thin, top=thin, bottom=thin)

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # Sheet 1: Executive Summary
    ws1 = wb.active
    ws1.title = 'Executive Summary'
    ws1.append(['รายงานผลการถอดแบบและประมาณราคาโครงสร้าง (Plan2BOQ Report)'])
    ws1.append([f'โครงการ: {project_name} (คุณนุชรินทร์ สิริใหม่) | สถานที่: ต.หนองยวง อ.เวียงหนองล่อง จ.ลำพูน'])
    ws1.append([])
    ws1.append(['ลำดับ', 'หมวดรายการวัสดุ', 'ปริมาณสุทธิ', 'เผื่อสูญเสีย (Waste)', 'ปริมาณรวมสั่งซื้อ', 'หน่วย', 'หมายเหตุตามแบบ 116-69'])

    summary_rows = [
        [1, 'คอนกรีตโครงสร้างหล่อในที่ 240 ksc (ฐานราก+ตอม่อ+เสา+คาน+พื้น S1)', 20.58, '+3%', 21.20, 'ลบ.ม. (คิว)', 'หล่อในที่ทุกหมวด'],
        [2, 'คอนกรีต Topping พื้นสำเร็จรูป PS (หนา 5 ซม.)', 4.71, '+3%', 4.85, 'ลบ.ม. (คิว)', 'เททับหน้าแผ่น PS'],
        [3, 'คอนกรีตหยาบรองก้นหลุมฐานราก (Lean 7.5cm)', 1.96, '-', 1.96, 'ลบ.ม. (คิว)', 'รองก้นหลุม 18 ฐาน'],
        [4, 'แผ่นพื้นสำเร็จรูป Hollow Core / Solid Plank (PS)', 94.20, '-', 95.00, 'ตร.ม.', 'สั่งผลิตตามความยาวช่วง'],
        [5, 'เหล็กเสริมรวม (DB12, RB9, RB6) — FFD', 1664.50, 'ตัดรวม FFD', 1664.50, 'กก. (224 เส้น)', 'ประหยัดได้ 26 เส้น (260ม.)'],
        [6, 'ลวดตะแกรงเหล็ก Wire Mesh 4mm @0.20m', 94.20, '+5%', 99.00, 'ตร.ม.', 'งานเท Topping'],
        [7, 'เหล็กรูปพรรณโครงสร้างหลังคา (2-C 150, C 100, แปกล่อง)', 2150.00, '+5%', 2257.50, 'กก. (2.26 ตัน)', 'อะเส/อกไก่/ดั้ง/แป'],
        [8, 'ไม้แบบหล่อคอนกรีตโครงสร้าง', 174.64, '-', 174.64, 'ตร.ม.', 'ฐานราก+ตอม่อ+เสา+คาน']
    ]
    for r in summary_rows:
        ws1.append(r)

    # Sheet 2: Footings
    ws2 = wb.create_sheet('1. ฐานราก (Footings)')
    ws2.append(['หมวดที่ 1: งานฐานราก ค.ส.ล. (Footings: F1, F2, F3) — รวม 18 ฐาน (แผ่น S-05)'])
    ws2.append([])
    ws2.append(['รหัส', 'จำนวนฐาน', 'กว้าง (ม.)', 'ยาว (ม.)', 'ลึก (ม.)', 'คอนกรีตสุทธิ (ม³)', 'คอนกรีตหยาบ (ม³)', 'เหล็กเสริมตะแกรง', 'จำนวนเหล็ก DB12 (เส้น 10ม.)', 'ไม้แบบ (ตร.ม.)'])
    footing_rows = [
        ['F1', 1, 0.75, 0.75, 0.20, 0.113, 0.048, 'DB12 @ 0.15 (4+4 เส้น)', 1, 0.60],
        ['F2', 10, 1.00, 1.00, 0.25, 2.500, 0.863, 'DB12 @ 0.15 (8+8 เส้น)', 14, 10.00],
        ['F3', 7, 1.30, 1.30, 0.30, 3.549, 1.045, 'DB12 @ 0.15 (8+8 เส้น)', 14, 10.92],
        ['รวมฐานราก', 18, '-', '-', '-', 6.162, 1.956, 'DB12 รวม 280.8 ม.', 29, 21.52]
    ]
    for r in footing_rows:
        ws2.append(r)

    # Sheet 3: Columns
    ws3 = wb.create_sheet('2. เสาและตอม่อ (Columns)')
    ws3.append(['หมวดที่ 2: งานตอม่อและเสา (Piers & Columns: C1) — รวม 18 ต้น (แผ่น S-05 & S-09)'])
    ws3.append([])
    ws3.append(['รหัสเสา', 'จำนวนต้น', 'หน้าตัด (ม.)', 'ความสูงตอม่อ (ม.)', 'ความสูงเสา (ม.)', 'คอนกรีตสุทธิ (ม³)', 'เหล็กยืน (S-09 Schedule)', 'เหล็กปลอก (S-09 Schedule)', 'ไม้แบบ (ตร.ม.)'])
    col_rows = [
        ['C1 บน F1', 1, '0.20 x 0.20', 1.20, 3.40, 0.184, '4-DB12 (L=5.10ม.)', '1-RB6 @ 0.15ม.', 3.68],
        ['C1 บน F2', 10, '0.20 x 0.20', 1.20, 3.40, 1.840, '4-DB12 (L=5.10ม.)', '1-RB6 @ 0.15ม.', 36.80],
        ['C1 บน F3', 7, '0.20 x 0.20', 1.20, 3.40, 1.288, '4-DB12 (L=5.10ม.)', '1-RB6 @ 0.15ม.', 25.76],
        ['รวมเสาและตอม่อ', 18, '-', '-', '-', 3.312, '4-DB12 (37 เส้น)', 'RB6 (42 เส้น)', 66.24]
    ]
    for r in col_rows:
        ws3.append(r)

    # Sheet 4: Beams
    ws4 = wb.create_sheet('3. คานคอดิน (Beams)')
    ws4.append(['หมวดที่ 3: งานคานคอดิน (Ground Beams: B1 - B4, CB) — แผ่น S-06 & S-10'])
    ws4.append([])
    ws4.append(['รหัสคาน', 'หน้าตัด (ม.)', 'จำนวนช่วง', 'ความยาวรวม (ม.)', 'คอนกรีต (ม³)', 'เหล็กบน (S-10 Schedule)', 'เหล็กล่าง (S-10 Schedule)', 'เหล็กปลอก', 'ไม้แบบ (ตร.ม.)'])
    beam_rows = [
        ['B1', '0.20 x 0.40', 12, 35.50, 2.840, '2-DB12 (Cont. 3-DB12)', '2-DB12', '1-RB6 @ 0.15ม.', 28.40],
        ['B2', '0.20 x 0.40', 4, 14.00, 1.120, '2-DB12 (Cont. 4-DB12)', '4-DB12', '1-RB6 @ 0.15ม.', 11.20],
        ['B3', '0.20 x 0.40', 6, 21.00, 1.680, '3-DB12 (Cont. 5-DB12)', '5-DB12', '1-RB6 @ 0.15ม.', 16.80],
        ['B4', '0.20 x 0.50', 8, 26.00, 2.600, '4-DB12 (Cont. 6-DB12)', '6-DB12', '1-RB6 @ 0.15ม.', 20.80],
        ['CB', '0.20 x 0.40', 4, 12.00, 0.960, '4-DB12', '2-DB12', '1-RB6 @ 0.15ม.', 9.60],
        ['รวมคานคอดิน', '-', 34, 108.50, 9.200, 'DB12: 84 เส้น', 'DB12', 'RB6: 98 เส้น', 86.80]
    ]
    for r in beam_rows:
        ws4.append(r)

    # Sheet 5: Floors
    ws5 = wb.create_sheet('4. งานพื้น (Floors)')
    ws5.append(['หมวดที่ 4: งานระบบพื้นอาคาร (Floor Slabs: PS, S1) — ผัง S-06 & สถาปัตย์'])
    ws5.append([])
    ws5.append(['รหัสพื้น', 'ประเภทระบบพื้น', 'พื้นที่สุทธิ (ตร.ม.)', 'ความหนา (ม.)', 'คอนกรีตสุทธิ (ม³)', 'เหล็กเสริม / ตะแกรง', 'หมายเหตุ'])
    floor_rows = [
        ['PS', 'แผ่นพื้นสำเร็จรูป Hollow Core / Solid Plank', 94.20, 0.05, 4.710, 'Wire Mesh 4mm @0.20m (94.20 ตร.ม.)', 'Topping หนา 5 ซม. (SFL=+0.95)'],
        ['S1', 'พื้นหล่อในที่ ค.ส.ล.', 24.30, 0.10, 2.430, 'RB9 @ 0.15ม. บน-ล่าง (238.4 กก.)', 'ห้องน้ำและระเบียง (SFL=+0.90)'],
        ['รวมงานพื้น', '-', 118.50, '-', 7.140, 'Topping 4.71 ม³ + เทในที่ 2.43 ม³', '-']
    ]
    for r in floor_rows:
        ws5.append(r)

    # Sheet 6: Roof
    ws6 = wb.create_sheet('5. โครงหลังคา (Roof)')
    ws6.append(['หมวดที่ 5: งานโครงสร้างหลังคาเหล็กรูปพรรณ (Roof Structure) — แผ่น S-07 & S-08'])
    ws6.append([])
    ws6.append(['ชิ้นส่วนโครงสร้าง', 'ขนาดหน้าตัดเหล็ก (มม.)', 'ระดับความสูง (El)', 'ความยาวรวม (ม.)', 'น้ำหนักสุทธิ (กก.)', 'หมายเหตุ'])
    roof_rows = [
        ['อะเสรอบอาคาร (Eaves Beam)', '2-C 150 x 50 x 20 x 3.2', 'El = +4.30 ม.', 112.40, 1013.8, 'เหล็ก C ประกบคู่ 2C'],
        ['อกไก่ / ดั้ง (Ridge & Posts)', '2-C 150 x 50 x 20 x 3.2', 'El = +5.19 ถึง +6.78 ม.', 54.60, 492.5, 'เหล็ก C ประกบคู่ 2C'],
        ['ตะเข้สัน / สะพานรับจันทัน', '2-C 150 x 50 x 20 x 3.2', 'El = varies', 38.20, 344.6, 'เหล็ก C ประกบคู่ 2C'],
        ['ระแนง / แปหลังคา', 'กล่อง 25 x 25 x 1.6', '@ 0.32 ม.', 282.00, 299.1, 'ระยะแปกระเบื้อง'],
        ['รวมโครงหลังคา', '-', '-', 487.20, 2150.0, 'รวมน้ำหนักเหล็กรูปพรรณ 2.15 ตัน']
    ]
    for r in roof_rows:
        ws6.append(r)

    # Sheet 7: Rebar Optimization FFD
    ws7 = wb.create_sheet('6. ตัดเศษเหล็ก (FFD)')
    ws7.append(['หมวดที่ 6: การบริหารเศษเหล็กเสริมรวมทั้งโครงการ (Bar Cutting List FFD Optimization)'])
    ws7.append([])
    ws7.append(['ขนาดเหล็กเสริม', 'ความยาวมาตรฐาน (ม.)', 'ตัดแยกหมวดเดิม (เส้น)', 'ตัดรวม FFD ใหม่ (เส้น)', 'ประหยัดสั่งซื้อ (เส้น)', 'ของเสีย (Waste %)'])
    rebar_rows = [
        ['เหล็กข้ออ้อย DB12', 10.0, 150, 134, 16, '1.4%'],
        ['เหล็กเส้นกลม RB9', 10.0, 28, 25, 3, '2.1%'],
        ['เหล็กเส้นกลม RB6', 10.0, 72, 65, 7, '1.1%'],
        ['รวมเหล็กเส้นทั้งโครงการ', 10.0, 250, 224, 26, 'ประหยัดต้นทุน ~4,600 บ.']
    ]
    for r in rebar_rows:
        ws7.append(r)

    # Styling for all sheets
    for ws in wb.worksheets:
        ws.views.sheetView[0].showGridLines = True
        ws.row_dimensions[1].height = 28
        ws.cell(row=1, column=1).font = font_title
        if ws.max_row >= 2:
            ws.cell(row=2, column=1).font = font_sub

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.row == 3 or (ws.title != 'Executive Summary' and cell.row == 3):
                    cell.font = font_header
                    cell.fill = fill_header_orange if ws.title == 'Executive Summary' else fill_header_dark
                    cell.alignment = align_center
                else:
                    cell.font = font_body
                    cell.alignment = align_left
                    if cell.row % 2 == 0:
                        cell.fill = fill_alt
                cell.border = border_cell

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    out_dir = os.path.join(BASE_DIR, '..', 'project', project_name, 'boq')
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f'Plan2BOQ_Complete_Structural_BOQ_{project_name}.xlsx')
    wb.save(out_path)
    print(f'Successfully generated Excel BOQ: {out_path}')

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--project-name', default='116-69 - แบบบ้านชั้นเดียว', help='Project name')
    args = parser.parse_args()
    generate_excel_boq(args.project_name)
